import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import re
from urllib.parse import urlparse


# Функция для чтения ссылок из файла
def read_urls_from_file(file_path):
    urls = []
    with open(file_path, 'r') as file:
        for line in file:
            url = line.strip()
            if url:  # Проверяем, что строка не пустая
                urls.append(get_counter_yadro_url(url))
    return urls


# Функция для получения и парсинга содержимого веб-страницы
def get_page_content(url):
    try:
        response = requests.get(url)
        response.raise_for_status()  # Проверяем, что запрос прошел успешно
        return response.text
    except requests.RequestException as e:
        print(f"Request failed: {e}")
        return None


# Функция для сохранения данных в Excel-файл
def save_to_excel(data, file_name='output.xlsx'):
    wb = Workbook()
    ws = wb.active
    
    # Записываем заголовки столбцов
    headers = ['URL', 'Месяц. Просмотры', 'Месяц. Посетители', 'Неделя. Просмотры', 'Неделя. Посетители', 'День. Просмотры', 'День. Посетители', 'Сегодня. Просмотры', 'Сегодня. Посетители', 'Просмотров онлайн', 'Посетителей онлайн']
    for col_num, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws['{}1'.format(col_letter)] = header
        ws.column_dimensions[col_letter].width = 15
    
    # Записываем данные
    for row_num, url_data in enumerate(data, start=2):
        ws[f'A{row_num}'] = url_data['LI_site_url']
        ws[f'B{row_num}'] = url_data['LI_month_hit']
        ws[f'C{row_num}'] = url_data['LI_month_vis']
        ws[f'D{row_num}'] = url_data['LI_week_hit']
        ws[f'E{row_num}'] = url_data['LI_week_vis']
        ws[f'F{row_num}'] = url_data['LI_day_hit']
        ws[f'G{row_num}'] = url_data['LI_day_vis']
        ws[f'H{row_num}'] = url_data['LI_today_hit']
        ws[f'I{row_num}'] = url_data['LI_today_vis']
        ws[f'J{row_num}'] = url_data['LI_online_hit']
        ws[f'K{row_num}'] = url_data['LI_online_vis']
    
    wb.save(filename=file_name)


def prev_data_format(str_data):
    pattern = r'(LI_\w+) = ([^\s;]+);'

    # Используем findall для поиска всех совпадений
    matches = re.findall(pattern, str_data.replace("'", '').replace('_site', '_site_url'))

    # Создаем словарь из найденных ключей и значений
    dictionary = dict(matches)
    return dictionary if dictionary else None

# Получаем ссылку на на статистику на counter.yadro.ru с нужным нам доменом
def get_counter_yadro_url(url):
    parsed_url = urlparse(url)
    netloc = parsed_url.netloc
    if netloc.startswith("www."):
        netloc = netloc[4:]  # Удаляем "www." из начала строки
    root_domain = netloc.split('.')[-2] + '.' + netloc.split('.')[-1]

    return f'https://counter.yadro.ru/values?site={root_domain}'


# Главная функция
def main(file_path, output_file_name='output.xlsx'):
    urls = read_urls_from_file(file_path)
    data = []
    
    for url in urls:
        content = get_page_content(url)
        if content:
            soup = BeautifulSoup(content, 'html.parser')
            page_info = prev_data_format(str(soup))
        
            if page_info:
                data.append(page_info)

    save_to_excel(data, output_file_name)


# Пример использования
file_path = 'urls.txt'  # Путь к файлу со списком URL
main(file_path)
